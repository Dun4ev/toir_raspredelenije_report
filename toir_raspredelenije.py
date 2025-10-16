import re
import shutil
from pathlib import Path
import sys
from datetime import datetime
import json
import os

SRC_DIR = Path(__file__).resolve().parent / "src"
if SRC_DIR.exists() and str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from toir_manager.core.logging_models import (  # noqa: E402
    TransferAction,
    TransferStatus,
)
from toir_manager.services.log_writer import DispatchLogger  # noqa: E402

# Для работы с Excel требуется установка библиотеки openpyxl: pip install openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "[КРИТИЧЕСКАЯ ОШИБКА] Библиотека openpyxl не найдена. Пожалуйста, установите ее: pip install openpyxl"
    )
    sys.exit(1)

# Настройка UTF-8 вывода в Windows-консоли
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined, union-attr]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined, union-attr]
except Exception:
    pass

# === НАСТРОЙКИ ПУТЕЙ ===
# Замените пути на актуальные для вашей системы.
# Используйте r"..." для путей в Windows, чтобы избежать проблем с \

# 1. Папка, где лежат исходные проекты для обработки (внутри которой папки с отчетами)
INBOX_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel\\00_Inbox"
)

# 2. Папка для копирования PDF-файлов (примечания)
NOTES_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel\\03_Notes"
)

# 3. Папка для копирования PDF-файлов (трансмиттал)
TRA_GST_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel\\04_TRA_GST"
)

# 4. Папка для специальной группировки
TRA_SUB_APP_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel\\05_TRA_SUB_app"
)

# 5. Корневая папка для создаваемой структуры (Год/Месяц/Part/pdf и Native)
DEST_ROOT_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel"
)

# 6. Временная папка для создания архивов (может совпадать с DEST_ROOT_DIR)
TEMP_ARCHIVE_DIR = Path(
    r"D:\\Code_and_Scripts_local\\_TEST_for\\toir_raspredelenije_test2_cel"
)


def _override_path(default: Path, env_name: str) -> Path:
    """Возвращает путь, переопределённый переменной окружения."""

    override = os.environ.get(env_name)
    if not override:
        return default
    candidate = Path(override).expanduser()
    try:
        return candidate.resolve()
    except OSError:
        return candidate


BOOL_TRUE_VALUES = {"1", "true", "yes", "on"}
BOOL_FALSE_VALUES = {"0", "false", "no", "off"}


PART_FILTER_DEFAULT = "CS/LP"
VALID_PART_FILTERS = {"LP", "CS", "CS/LP"}

PERIOD_TRANSLATION = {
    "С": "C",
}


def _env_flag(name: str, default: bool) -> bool:
    """Считывает булеву переменную окружения с резервным значением."""

    value = os.environ.get(name)
    if value is None:
        return default
    normalized = value.strip().lower()
    if normalized in BOOL_TRUE_VALUES:
        return True
    if normalized in BOOL_FALSE_VALUES:
        return False
    return default


def _get_part_filter() -> str:
    """Возвращает фильтр по полю part (CS/LP)."""

    raw = os.environ.get("TOIR_PART_FILTER")
    if not raw:
        return PART_FILTER_DEFAULT
    normalized = raw.strip().upper()
    if normalized not in VALID_PART_FILTERS:
        print(
            f"[WARN] Неподдерживаемое значение TOIR_PART_FILTER={raw}; используется CS/LP по умолчанию."
        )
        return PART_FILTER_DEFAULT
    return normalized


INBOX_DIR = _override_path(INBOX_DIR, "TOIR_INBOX_DIR")
NOTES_DIR = _override_path(NOTES_DIR, "TOIR_NOTES_DIR")
TRA_GST_DIR = _override_path(TRA_GST_DIR, "TOIR_TRA_GST_DIR")
TRA_SUB_APP_DIR = _override_path(TRA_SUB_APP_DIR, "TOIR_TRA_SUB_APP_DIR")
DEST_ROOT_DIR = _override_path(DEST_ROOT_DIR, "TOIR_DEST_ROOT_DIR")
TEMP_ARCHIVE_DIR = _override_path(TEMP_ARCHIVE_DIR, "TOIR_TEMP_ARCHIVE_DIR")

# 7. Путь к файлу-справочнику
TZ_FILE_PATH = Path("Template/TZ_glob.xlsx")


# === НАСТРОЙКИ ЛОГИКИ ===

# Регулярное выражение для разбора имени файла на основе предоставленной схемы
RE_FILENAME = re.compile(
    r"""^CT-
    (?P<type>CL|DR)-
    (?P<scope>[A-Z0-9]+)-
    (?P<part>CS|LP)-
    (?P<object_name>[A-Z0-9]+)-
    (?P<tz_index>[\w\.]+)-
    (?P<reserved>\d{2})-
    (?P<period>\w+)-
    (?P<date>\d{8})-
    (?P<revision>\d{2})
    _All.*?\.pdf$""",  # Ищем _All, затем любые символы, и только .pdf
    re.IGNORECASE | re.VERBOSE,
)


# Словарь для перевода номера месяца в название
MONTH_MAP = {
    "01": "January",
    "02": "February",
    "03": "March",
    "04": "April",
    "05": "May",
    "06": "June",
    "07": "July",
    "08": "August",
    "09": "September",
    "10": "October",
    "11": "November",
    "12": "December",
}

# Настройки для файла-справочника TZ_glob.xlsx
TZ_SHEET_NAME = "gen_cl"
TZ_LOOKUP_COL = "B"  # Колонка с индексами (I.7.5)
TZ_SUFFIX_COL = "G"

CS_FOLDER_OVERRIDES: dict[str, str] = {
    "II.1.1": "II.1",
    "II.1.2": "II.1",
    "II.1.3": "II.1",
    "II.1.4": "II.1",
    "II.1.5": "II.1",
    "II.1.6": "II.1",
    "II.1.7": "II.1",
    "II.1.8": "II.1",
    "II.2.1": "II.2",
    "II.2.2": "II.2",
    "II.2.3": "II.2",
    "II.2.4": "II.2",
    "II.2.5": "II.2",
    "II.2.6": "II.2",
    "II.2.7": "II.2",
    "II.2.8": "II.2",
    "II.2.9": "II.2",
    "II.2.10": "II.2",
    "II.2.11": "II.2",
    "II.2.12": "II.2",
    "II.2.13": "II.2",
    "II.3.1": "II.3",
    "II.3.2": "II.3",
    "II.3.3": "II.3",
    "II.3.4": "II.3",
    "II.3.5": "II.3",
    "II.4.1": "II.4",
    "II.5.1": "II.5",
    "II.6.1": "II.6",
    "II.6.2": "II.6",
    "II.7.1": "II.7",
    "II.8.1": "II.8",
    "II.8.1.1": "II.8",
    "II.8.2": "II.8",
    "II.8.2.1": "II.8",
    "II.8.3": "II.8",
    "II.8.3.1": "II.8",
    "II.8.4": "II.8",
    "II.8.5": "II.8",
    "II.8.6": "II.8",
    "II.8.7": "II.8",
    "II.9.1": "II.9",
    "II.18.2": "II.18",
    "II.19.1": "II.19",
    "II.19.2": "II.19",
    "II.22.1": "II.22",
    "II.22.3": "II.22",
    "II.22.4": "II.22",
    "II.23.1": "II.23",
    "II.23.2": "II.23",
    "II.23.3": "II.23",
    "II.25.1": "II.25",
    "II.26.1.1": "II.26",
    "II.26.2": "II.26",
    "II.26.3": "II.26",
    "II.26.4": "II.26",
    "II.27.1": "II.27",
    "II.28.1": "II.28",
    "II.29.2": "II.29",
    # Добавляйте остальные индексы по мере необходимости
}

CS_DEFAULT_FOLDERS: dict[str, str] = {
    "II.2": "Корректирующее обслуживание",
    "II.18": "II.18_UPS",
}


LP_FOLDER_OVERRIDES: dict[str, str] = {
    # Добавляйте сопоставления, если объект должен попадать в другую папку
}


LOGGER: DispatchLogger | None = None


def _merge_metadata(
    base: dict[str, str] | None, extra: dict[str, str]
) -> dict[str, str]:
    """Скомбинировать базовые и дополнительные метаданные."""

    result: dict[str, str] = {} if base is None else dict(base)
    for key, value in extra.items():
        if value is not None:
            result[key] = value
    return result


def _log_success(
    action: TransferAction,
    source: Path,
    target: Path | None,
    metadata: dict[str, str] | None = None,
) -> None:
    """Безопасно записать успешную операцию в журнал."""

    if LOGGER is None:
        return
    try:
        LOGGER.log_success(
            action=action, source_path=source, target_path=target, metadata=metadata
        )
    except Exception:
        pass


def _log_error(
    action: TransferAction,
    source: Path,
    target: Path | None,
    message: str,
    metadata: dict[str, str] | None = None,
) -> None:
    """Безопасно записать ошибку операции."""

    if LOGGER is None:
        return
    try:
        LOGGER.log_error(
            action=action,
            source_path=source,
            target_path=target,
            message=message,
            metadata=metadata,
        )
    except Exception:
        pass


def _log_destination_event(
    event: str,
    source: Path,
    target: Path,
    metadata: dict[str, str] | None = None,
) -> None:
    """Логирует событие для каталога назначения."""

    if LOGGER is None:
        return
    try:
        LOGGER.log(
            action=TransferAction.COPY_DESTINATION,
            status=TransferStatus.SUCCESS,
            source_path=source,
            target_path=target,
            message=event,
            metadata=_merge_metadata(metadata, {"destination_event": event}),
        )
    except Exception:
        pass


# Колонка с суффиксами (краткая аббревиатура)


def find_suffix_in_tz_file(lookup_key: str) -> str | None:
    """
    Ищет индекс в файле TZ_glob.xlsx и возвращает суффикс.
    """
    if not TZ_FILE_PATH.exists():
        print(f"  - [ОШИБКА] Файл-справочник не найден: {TZ_FILE_PATH}")
        return None

    try:
        wb = load_workbook(TZ_FILE_PATH, data_only=True)
        if TZ_SHEET_NAME not in wb.sheetnames:
            print(
                f"  - [ОШИБКА] Лист '{TZ_SHEET_NAME}' не найден в файле {TZ_FILE_PATH}"
            )
            return None

        ws = wb[TZ_SHEET_NAME]

        lookup_col_idx = ord(TZ_LOOKUP_COL.upper()) - ord("A") + 1
        suffix_col_idx = ord(TZ_SUFFIX_COL.upper()) - ord("A") + 1

        for row in ws.iter_rows(
            min_row=2, values_only=True
        ):  # Начинаем со второй строки, пропуская заголовок
            cell_value = (
                str(row[lookup_col_idx - 1]).strip().lower()
                if row[lookup_col_idx - 1]
                else ""
            )
            if cell_value == lookup_key.strip().lower():
                suffix = row[suffix_col_idx - 1]
                return str(suffix).strip() if suffix else None

        return None
    except Exception as e:
        print(f"  - [ОШИБКА] Ошибка при чтении файла {TZ_FILE_PATH}: {e}")
        return None


def process_special_grouping_for_sub_app(
    report_file: Path, data: dict, metadata: dict[str, str] | None = None
) -> None:
    """Организовать дополнительную выгрузку файла в каталог 05_TRA_SUB_app."""

    print(f"  - Обрабатываем дополнительную группировку для {TRA_SUB_APP_DIR.name}...")
    try:
        grouping_key = f"{data['tz_index']}-{data['reserved']}-{data['period']}"
        period_raw = data["period"].upper()
        period = PERIOD_TRANSLATION.get(period_raw, period_raw)

        folder_name = ""
        extra_metadata = _merge_metadata(
            metadata,
            {
                "grouping_key": grouping_key,
            },
        )

        if period == "C":  # Обрабатываем латинский и кириллический варианты
            folder_name = grouping_key
            extra_metadata = _merge_metadata(
                extra_metadata,
                {
                    "tra_sub_folder": folder_name,
                },
            )
            print(f"    - Режим 'еженедельный'. Итоговая папка: {folder_name}")
        else:
            tz_index = data["tz_index"]
            print(f"    - Ищем суффикс для узла: {tz_index}")
            suffix = find_suffix_in_tz_file(tz_index)

            if not suffix:
                message = f"Не найден суффикс в TZ_glob.xlsx для '{tz_index}'"
                _log_error(
                    TransferAction.COPY_TRA_SUB,
                    report_file,
                    None,
                    message,
                    _merge_metadata(
                        extra_metadata,
                        {
                            "tz_index": tz_index,
                            "reason": "missing_suffix",
                        },
                    ),
                )
                return

            print(f"    - Используем суффикс: '{suffix}'")
            folder_name = f"{grouping_key}_{suffix}"
            extra_metadata = _merge_metadata(
                extra_metadata,
                {
                    "tra_sub_folder": folder_name,
                    "tz_suffix": suffix,
                },
            )

        base_dest_dir = TRA_SUB_APP_DIR
        date_str = data.get("date")
        year_dir: str | None = None
        month_folder_name: str | None = None

        if date_str:
            parsed_date: datetime | None = None
            try:
                parsed_date = datetime.strptime(date_str, "%Y%m%d")
            except (ValueError, TypeError):
                parsed_date = None

            if parsed_date is not None:
                year_dir = parsed_date.strftime("%Y")
                month_num = parsed_date.strftime("%m")
                month_name = MONTH_MAP.get(month_num, "UnknownMonth")
                month_folder_name = f"{month_num}.{month_name}"

        if year_dir and month_folder_name:
            base_dest_dir = base_dest_dir / year_dir / month_folder_name
            extra_metadata = _merge_metadata(
                extra_metadata,
                {
                    "tra_sub_year": year_dir,
                    "tra_sub_month": month_folder_name,
                },
            )

        dest_dir = base_dest_dir / folder_name
        dest_dir.mkdir(parents=True, exist_ok=True)
        print(f"    - Копируем отчёт в каталог: {dest_dir}")
        shutil.copy(report_file, dest_dir)
        _log_success(
            TransferAction.COPY_TRA_SUB,
            report_file,
            dest_dir / report_file.name,
            extra_metadata,
        )

    except Exception as e:  # noqa: BLE001
        message = f"Ошибка копирования в каталог TRA_SUB_app: {e}"
        _log_error(
            TransferAction.COPY_TRA_SUB,
            report_file,
            None,
            message,
            _merge_metadata(
                metadata,
                {
                    "grouping_key": grouping_key,
                    "tra_sub_folder": folder_name or grouping_key,
                },
            ),
        )


def normalize_object_name(object_name: str) -> str:
    """
    Нормализует имя объекта, удаляя ведущий ноль для однозначных номеров.
    Пример: BVS05 -> BVS5. BVS10 -> BVS10.
    """
    # Ищем шаблон: BVS, затем 0, затем одна цифра от 1 до 9
    match = re.match(r"^(BVS)0([1-9])$", object_name, re.IGNORECASE)
    if match:
        # Собираем новое имя из первой группы (BVS) и второй (цифра)
        normalized_name = match.group(1) + match.group(2)
        print(
            f"  - [ИНФО] Имя объекта нормализовано: {object_name} -> {normalized_name}"
        )
        return normalized_name
    return object_name


def copy_to_gst_folder(
    report_file: Path,
    date_str: str,
    tra_gst_dir: Path,
    metadata: dict[str, str] | None = None,
) -> None:
    """Разложить отчёт в каталог 04_TRA_GST по рабочим неделям."""

    print(f"  - Проверяем папку {tra_gst_dir.name} для распределения...")

    try:
        date_obj = datetime.strptime(date_str, "%Y%m%d")
        week_number = date_obj.isocalendar()[1]
        year = date_str[:4]
    except ValueError:
        message = f"Некорректная дата в имени файла: '{date_str}'."
        print(f"  - [Ошибка] {message}")
        _log_error(TransferAction.COPY_GST, report_file, None, message, metadata)
        return

    while True:
        folder_name = f"{year}_T{week_number}_GST"
        target_dir = tra_gst_dir / folder_name

        print(f"    - Целевая папка: {target_dir.name}")

        is_locked = False
        if target_dir.exists():
            for ext in ["*.zip", "*.7z", "*.rar"]:
                if any(
                    "CT-GST-TRA-PRM-" in archive.name
                    for archive in target_dir.glob(ext)
                ):
                    print(
                        "    - Обнаружены архивы в каталоге, подбираем следующую неделю..."
                    )
                    is_locked = True
                    break

        if is_locked:
            week_number += 1
            continue

        print("    - Папка свободна, копируем отчёт...")
        try:
            target_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy(report_file, target_dir)
            extra_metadata = _merge_metadata(
                metadata,
                {
                    "gst_folder": target_dir.name,
                    "week": str(week_number),
                },
            )
            _log_success(
                TransferAction.COPY_GST,
                report_file,
                target_dir / report_file.name,
                extra_metadata,
            )
            print(f"    - Файл успешно помещён в {target_dir}")
            break
        except Exception as e:  # noqa: BLE001
            message = f"Ошибка копирования в каталог GST: {e}"
            print(f"  - [Ошибка] {message}")
            _log_error(
                TransferAction.COPY_GST,
                report_file,
                target_dir / report_file.name,
                message,
                _merge_metadata(
                    metadata,
                    {
                        "gst_folder": target_dir.name,
                        "week": str(week_number),
                    },
                ),
            )
            break


def find_project_folders(inbox_dir: Path) -> list[Path]:
    """Рекурсивно находит каталоги, содержащие файлы `_All`."""

    project_dirs: set[Path] = set()
    for pdf_path in inbox_dir.rglob("*_All*.[pP][dD][fF]"):
        if pdf_path.is_file():
            project_dirs.add(pdf_path.parent)
    return sorted(project_dirs, key=lambda item: item.as_posix().lower())


def process_project_folder(project_path: Path) -> None:
    """Обработать проектную папку из INBOX."""
    print(f"\n--- Обрабатываем проект: {project_path.name} ---")

    all_matching_files = []
    for file_path in project_path.glob("*_All*.[pP][dD][fF]"):
        if RE_FILENAME.match(file_path.name):
            all_matching_files.append(file_path)

    if not all_matching_files:
        print("  - [Предупреждение] Подходящих файлов не найдено. Пропускаем.")
        return
    if len(all_matching_files) > 1:
        print(
            f"  - [Внимание] Найдено несколько файлов ({len(all_matching_files)}). Берём первый."
        )

    report_file = all_matching_files[0]
    print(f"  - Выбран файл: {report_file.name}")

    match = RE_FILENAME.match(report_file.name)
    if not match:
        print(
            f"  - [Ошибка] Имя файла {report_file.name} не соответствует шаблону. Пропускаем."
        )
        return

    data = match.groupdict()
    attributes_dump = json.dumps(data, indent=4, ensure_ascii=False)
    print("  - Извлечённые атрибуты:")
    print(attributes_dump)

    date_str = data["date"]
    year = date_str[:4]
    month_num = date_str[4:6]
    month_name = MONTH_MAP.get(month_num, "UnknownMonth")
    part = data["part"].upper()
    part_filter = _get_part_filter()
    if part_filter != "CS/LP" and part != part_filter:
        print(
            f"  - [INFO] Пропуск из-за фильтра part: {part} не входит в {part_filter}."
        )
        return

    month_folder_name = f"{month_num}.{month_name}"
    period_raw = data["period"].upper()
    period = PERIOD_TRANSLATION.get(period_raw, period_raw)

    base_metadata = {
        k: v
        for k, v in {
            "project_folder": project_path.name,
            "type": data.get("type"),
            "scope": data.get("scope"),
            "part": part,
            "object_name": data.get("object_name"),
            "tz_index": data.get("tz_index"),
            "period": period,
        }.items()
        if v
    }

    notes_enabled = _env_flag("TOIR_ENABLE_NOTES", True)
    tra_gst_enabled = _env_flag("TOIR_ENABLE_TRA_GST", True)
    tra_sub_app_enabled = _env_flag("TOIR_ENABLE_TRA_SUB_APP", True)
    dest_root_enabled = _env_flag("TOIR_ENABLE_DEST_ROOT", True)

    pdf_dest_dir: Path | None = None
    archive_dest_dir: Path | None = None

    if dest_root_enabled:
        if period == "C":
            print("  - [Инфо] Рабочий режим: еженедельный (C).")
            target_folder_name = "Корректирующее обслуживание"
            pdf_dest_dir = (
                DEST_ROOT_DIR
                / year
                / month_folder_name
                / part
                / "pdf"
                / target_folder_name
            )
            archive_dest_dir = (
                DEST_ROOT_DIR
                / year
                / month_folder_name
                / part
                / "Native"
                / target_folder_name
            )
        else:
            print("  - [Инфо] Рабочий режим: стандартный.")
            if part == "LP":
                print("  - [Инфо] Раздел LP.")
                object_name_raw = data["object_name"].upper()
                object_name = normalize_object_name(object_name_raw)
                folder_name = LP_FOLDER_OVERRIDES.get(object_name, object_name)
                if folder_name != object_name:
                    print(
                        f"  - [Инфо] Используем сопоставление LP: {object_name} → {folder_name}"
                    )
                pdf_dest_dir = (
                    DEST_ROOT_DIR
                    / year
                    / month_folder_name
                    / part
                    / "pdf"
                    / folder_name
                )
                archive_dest_dir = (
                    DEST_ROOT_DIR
                    / year
                    / month_folder_name
                    / part
                    / "Native"
                    / folder_name
                )
                base_metadata = _merge_metadata(
                    base_metadata,
                    {
                        "destination_folder": folder_name,
                        "destination_prefix": folder_name,
                    },
                )
            elif part == "CS":
                print("  - [Инфо] Раздел CS.")
                tz_index = data["tz_index"]
                base_metadata = _merge_metadata(base_metadata, {"tz_index": tz_index})

                folder_prefix = CS_FOLDER_OVERRIDES.get(tz_index, tz_index)
                if folder_prefix != tz_index:
                    print(
                        f"  - [Инфо] Используем префикс из справочника: {folder_prefix}"
                    )

                pdf_parent = DEST_ROOT_DIR / year / month_folder_name / part / "pdf"
                native_parent = (
                    DEST_ROOT_DIR / year / month_folder_name / part / "Native"
                )
                pdf_parent.mkdir(parents=True, exist_ok=True)
                native_parent.mkdir(parents=True, exist_ok=True)

                found_folders = sorted(pdf_parent.glob(f"{folder_prefix}*"))
                destination_event = "found"
                if found_folders:
                    target_folder_name = found_folders[0].name
                    if len(found_folders) > 1:
                        print(
                            f"  - [Внимание] Несколько совпадений, берём {target_folder_name}"
                        )
                    pdf_dest_dir = found_folders[0]
                else:
                    target_folder_name = CS_DEFAULT_FOLDERS.get(
                        folder_prefix, folder_prefix
                    )
                    pdf_dest_dir = pdf_parent / target_folder_name
                    pdf_dest_dir.mkdir(parents=True, exist_ok=True)
                    destination_event = "created"
                    print(f"  - [Инфо] Создаём каталог: {pdf_dest_dir}")

                archive_dest_dir = native_parent / target_folder_name
                archive_dest_dir.mkdir(parents=True, exist_ok=True)

                base_metadata = _merge_metadata(
                    base_metadata,
                    {
                        "destination_folder": target_folder_name,
                        "destination_prefix": folder_prefix,
                        "destination_path": str(pdf_dest_dir),
                    },
                )
                _log_destination_event(
                    destination_event,
                    report_file,
                    pdf_dest_dir,
                    base_metadata,
                )
    else:
        print("  - [INFO] Skipping DEST_ROOT distribution due to settings.")
        return

    if dest_root_enabled and (pdf_dest_dir is None or archive_dest_dir is None):
        message = "Не удалось определить директорию назначения."
        print(f"  - [Ошибка] {message}")
        _log_error(
            TransferAction.COPY_DESTINATION, report_file, None, message, base_metadata
        )
        return

    assert pdf_dest_dir is not None
    assert archive_dest_dir is not None

    if notes_enabled:
        try:
            notes_target = NOTES_DIR / report_file.name
            shutil.copy(report_file, NOTES_DIR)
            _log_success(
                TransferAction.COPY_NOTES,
                report_file,
                notes_target,
                _merge_metadata(base_metadata, {"notes_dir": str(NOTES_DIR)}),
            )
            print(f"  - File copied to {NOTES_DIR}")
        except Exception as e:  # noqa: BLE001
            message = f"Failed to copy to {NOTES_DIR}: {e}"
            print(f"  - [ERROR] {message}")
            _log_error(
                TransferAction.COPY_NOTES,
                report_file,
                notes_target,
                message,
                base_metadata,
            )
            return
    else:
        print("  - [INFO] NOTES distribution disabled by settings.")

    if tra_gst_enabled:
        copy_to_gst_folder(report_file, date_str, TRA_GST_DIR, metadata=base_metadata)
    else:
        print("  - [INFO] TRA_GST distribution disabled by settings.")
    if tra_sub_app_enabled:
        process_special_grouping_for_sub_app(report_file, data, metadata=base_metadata)
    else:
        print("  - [INFO] 05_TRA_SUB_app distribution disabled by settings.")

    try:
        pdf_dest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy(report_file, pdf_dest_dir)
        pdf_target = pdf_dest_dir / report_file.name
        _log_success(
            TransferAction.COPY_DESTINATION,
            report_file,
            pdf_target,
            _merge_metadata(base_metadata, {"destination_path": str(pdf_dest_dir)}),
        )
        print(f"  - Файл скопирован в {pdf_dest_dir}")
    except Exception as e:  # noqa: BLE001
        message = f"Ошибка копирования в каталог назначения: {e}"
        print(f"  - [Ошибка] {message}")
        _log_error(
            TransferAction.COPY_DESTINATION,
            report_file,
            pdf_dest_dir / report_file.name,
            message,
            base_metadata,
        )
        return

    archive_target_path = (
        DEST_ROOT_DIR
        / year
        / month_folder_name
        / part
        / "Native"
        / f"{project_path.name}.zip"
    )
    try:
        archive_dest_dir.mkdir(parents=True, exist_ok=True)
        archive_basename = TEMP_ARCHIVE_DIR / project_path.name
        print(f"  - Создаём архив для каталога: {project_path.name}...")
        archive_path_str = shutil.make_archive(
            str(archive_basename), "zip", str(project_path)
        )
        archive_path = Path(archive_path_str)
        _log_success(
            TransferAction.CREATE_ARCHIVE,
            project_path,
            archive_path,
            _merge_metadata(base_metadata, {"archive_tmp": str(archive_path)}),
        )
        print(f"  - Копируем архив в: {archive_dest_dir}")
        shutil.copy(archive_path, archive_dest_dir)
        _log_success(
            TransferAction.COPY_ARCHIVE,
            archive_path,
            archive_dest_dir / archive_path.name,
            _merge_metadata(base_metadata, {"archive_dest": str(archive_dest_dir)}),
        )
        archive_path.unlink()
    except Exception as e:  # noqa: BLE001
        message = f"Ошибка обработки архива: {e}"
        print(f"  - [Ошибка] {message}")
        _log_error(
            TransferAction.COPY_ARCHIVE,
            project_path,
            archive_target_path,
            message,
            base_metadata,
        )


def main(inbox_dir: Path | None = None) -> None:
    """Точка входа обработки PDF."""
    global LOGGER

    target_inbox = Path(inbox_dir).resolve() if inbox_dir else INBOX_DIR

    print("Запуск распределения PDF...")
    with DispatchLogger() as logger:
        LOGGER = logger
        print(f"Текущий лог доступен в: {logger.file_path}")
        try:
            for dir_path in [NOTES_DIR, TRA_GST_DIR]:
                if not dir_path.exists():
                    print(f"Создаём вспомогательную директорию: {dir_path}")
                    dir_path.mkdir(parents=True, exist_ok=True)

            if not target_inbox.exists():
                print(f"[Ошибка] Входной каталог отсутствует: {target_inbox}")
                return
            stray_pdfs = [
                item
                for item in target_inbox.iterdir()
                if item.is_file() and item.suffix.lower() == ".pdf"
            ]
            if stray_pdfs:
                print(
                    "[Предупреждение] В корне входного каталога обнаружены PDF-файлы. "
                    "Каждый отчёт должен лежать в отдельной папке. Обработка остановлена."
                )
                for pdf_path in stray_pdfs:
                    _log_error(
                        TransferAction.COPY_DESTINATION,
                        pdf_path,
                        None,
                        "Файл расположен в корне INBOX. Требуется отдельная папка для каждого отчёта.",
                    )
                return

            project_folders = find_project_folders(target_inbox)
            if not project_folders:
                print(f"В {target_inbox} не найдено файлов `_All` для обработки.")
                return

            print(f"Найдено {len(project_folders)} папок с `_All` в {target_inbox}.")
            for folder in project_folders:
                process_project_folder(folder)

            print("\nОбработка завершена.")
        finally:
            LOGGER = None


if __name__ == "__main__":
    override = os.environ.get("TOIR_INBOX_DIR")
    override_path = Path(override).resolve() if override else None
    main(inbox_dir=override_path)
if __name__ == "__main__":
    main()
